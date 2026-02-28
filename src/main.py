import pandas as pd
import sys
from transform import time_complexity_to_df
from extract import read_input, parse_row, parse_report, parse_report_transactions
from config import FILE_PATH, FILE_PATH_TRANSACTIONS
from openpyxl import load_workbook
import cProfile
import pstats
import re

# wb = load_workbook(FILE_PATH)
# ws = wb.active

# for merged in list(ws.merged_cells.ranges):
#     min_col, min_row, max_col, max_row = merged.bounds
#     value = ws.cell(min_row, min_col).value
#     ws.unmerge_cells(str(merged))

#     # wypełnij wszystkie komórki
#     for r in range(min_row, max_row + 1):
#         for c in range(min_col, max_col + 1):
#             ws.cell(r, c).value = value

# wb.save("normalized.xlsx")

# df = pd.read_excel("normalized.xlsx")

def main():
    
    def save_raport(df, OUTPUT_PATH):
         df.to_excel(OUTPUT_PATH, index=False)
   
    df_dane = parse_report(FILE_PATH)
    df_dane_transactions = parse_report_transactions(FILE_PATH_TRANSACTIONS)
    
    df_dane_transactions = df_dane_transactions.copy()
    
    KEY_COL = "description"   # kolumna z nazwą produktu
    KEY_COL_trans = "Description"
    ITEM_COL = "item"

    #dopasowanie

    mapping = (
    df_dane
    .drop_duplicates(subset=KEY_COL)
    .set_index(KEY_COL)[ITEM_COL]
)

    df_dane_transactions[ITEM_COL] = df_dane_transactions[KEY_COL_trans].map(mapping)
    missing = df_dane_transactions[ITEM_COL].isna().sum()
    print("Brak dopasowań:", missing)

    QTY_COL = 5
    VALUE_COL = 6
    PRICE_COL = 7
    Count_COL = 3

    qty = pd.to_numeric(df_dane_transactions.iloc[:, QTY_COL], errors="coerce").fillna(1)
    count = pd.to_numeric(df_dane.iloc[:, Count_COL], errors="coerce").fillna(1)

# maska do powielenia
    mask_expand = qty.abs() > 1 # .abs() to metoda Pandas (i NumPy), która zwraca wartość bezwzględną liczby.
    mask_expand_dane = count > 1
# maska do pozostawienia wierszy bez zmian (1 lub -1)
    mask_keep = qty.abs() == 1
    mask_keep_dane = count.isin([1])
# część do powielenia
    df_expand = df_dane_transactions.loc[mask_expand].copy()
    qty_expand = qty[mask_expand]

    df_expand = df_expand.loc[df_expand.index.repeat(qty_expand.abs())].copy()

    # znak (+1 lub -1)
    sign = qty_expand.apply(lambda x: 1 if x > 0 else -1)

    # powiel znak do długości expand
    sign_repeated = sign.repeat(qty_expand.abs()).values

    df_expand_dane = df_dane.loc[mask_expand_dane].copy()
    qty_expand_dane = count[mask_expand_dane].astype(int)

    df_expand_dane = df_expand_dane.loc[df_expand_dane.index.repeat(qty_expand_dane)].copy()

# podział wartości i zaokrąglenie

    # df_expand.iloc[:, VALUE_COL] = (
    #     df_expand.iloc[:, VALUE_COL] / qty_expand.repeat(qty_expand)
    # ).round(2)
    # df_expand.iloc[:, PRICE_COL] = (
    #     df_expand.iloc[:, PRICE_COL] / qty_expand.repeat(qty_expand)
    # ).round(2)
    
    df_expand.iloc[:, QTY_COL] = sign_repeated
    df_expand_dane.iloc[:, Count_COL] = 1
# część bez zmian
    df_keep = df_dane_transactions.loc[mask_keep].copy()
    df_keep__dane = df_dane.loc[mask_keep_dane].copy()
# scalanie
    df_final = pd.concat([df_keep, df_expand]).sort_index().reset_index(drop=True)
    #print(df_final)  
    df_final_dane = pd.concat([df_keep__dane, df_expand_dane]).sort_index().reset_index(drop=True)
    df_final_dane = df_final_dane.drop(df_final_dane.columns[[1,4,5,6,7]], axis=1)
    
    
#     df_expanded = df_expanded.reset_index(drop=True)
    df_final = df_final.drop(df_final.columns[[0,1,3,4,6,7,8,9,10,11,12]], axis=1)
    col = df_final.pop(ITEM_COL)
    df_final.insert(0, ITEM_COL, col)
    
    #save_raport(df_final, "final.xlsx")
    df_dane = df_dane.drop(df_dane.columns[[1,4,5,6,7]], axis=1)
    df_final_renamed = df_final.rename(columns={
    "Description": "description",
    "Qty": "count"
})
    df_final2 = pd.concat([df_final_dane, df_final_renamed], ignore_index=True)
    save_raport(df_final2, "final.xlsx")
    print(df_final2)
    
if __name__ == "__main__":

    main()